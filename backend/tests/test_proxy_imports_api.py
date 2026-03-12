"""
Tests for PX-2A: proxy nomination and financial position import endpoints.

POST /api/admin/buildings/{id}/lot-owners/import-proxies
POST /api/admin/buildings/{id}/lot-owners/import-financial-positions

Structure:
  # --- Happy path ---
  # --- Input validation ---
  # --- Boundary values ---
  # --- State / precondition errors ---
  # --- Edge cases ---
"""
from __future__ import annotations

import csv
import io
import uuid

import openpyxl
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Building,
    FinancialPosition,
    LotOwner,
    LotProxy,
)
from app.models.lot_owner_email import LotOwnerEmail


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode()


def make_excel(headers: list, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def client(app):
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as ac:
        yield ac


@pytest_asyncio.fixture
async def building(db_session: AsyncSession) -> Building:
    b = Building(name="Proxy Test Building", manager_email="proxy@test.com")
    db_session.add(b)
    await db_session.flush()
    await db_session.refresh(b)
    return b


@pytest_asyncio.fixture
async def building_with_owners(db_session: AsyncSession) -> Building:
    b = Building(name="Proxy Owners Building", manager_email="po@test.com")
    db_session.add(b)
    await db_session.flush()

    lo1 = LotOwner(building_id=b.id, lot_number="1A", unit_entitlement=100)
    lo2 = LotOwner(building_id=b.id, lot_number="2B", unit_entitlement=50)
    lo3 = LotOwner(building_id=b.id, lot_number="3C", unit_entitlement=75)
    db_session.add_all([lo1, lo2, lo3])
    await db_session.flush()

    db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="owner1@test.com"))
    db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="owner2@test.com"))
    await db_session.flush()
    await db_session.refresh(b)
    return b


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-proxies (CSV)
# ---------------------------------------------------------------------------


class TestImportProxiesCSV:
    # --- Happy path ---

    async def test_upserts_new_proxy_nominations(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 2
        assert data["removed"] == 0
        assert data["skipped"] == 0

        # Verify DB
        result = await db_session.execute(
            select(LotProxy).join(LotOwner).where(LotOwner.building_id == building_with_owners.id)
        )
        proxies = result.scalars().all()
        assert len(proxies) == 2
        emails = {p.proxy_email for p in proxies}
        assert "proxy1@test.com" in emails
        assert "proxy2@test.com" in emails

    async def test_updates_existing_proxy_nomination(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # First upload
        csv_data1 = make_csv(["Lot#", "Proxy Email"], [["1A", "old_proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Second upload with updated email
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", "new_proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 1
        assert data["removed"] == 0

    async def test_blank_proxy_email_removes_nomination(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Create proxy first
        csv_data1 = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Now remove by blank email
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 1
        assert data["skipped"] == 0

    async def test_blank_proxy_email_no_existing_proxy_is_noop(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_unknown_lot_number_skipped(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy@test.com"], ["999X", "ghost@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 1
        assert data["skipped"] == 1

    async def test_lots_not_in_file_are_unaffected(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Set proxy for both 1A and 2B
        csv_data1 = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Second upload only touches 1A
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", "updated@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )

        # 2B proxy should still exist
        result = await db_session.execute(
            select(LotProxy)
            .join(LotOwner)
            .where(LotOwner.building_id == building_with_owners.id)
        )
        proxies = result.scalars().all()
        emails = {p.proxy_email for p in proxies}
        assert "proxy2@test.com" in emails

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email", "Extra"],
            [["1A", "proxy@test.com", "ignore_me"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    async def test_empty_csv_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_case_insensitive_headers(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["LOT#", "PROXY EMAIL"],
            [["1A", "proxy@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    # --- Input validation ---

    async def test_missing_lot_hash_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Proxy Email"], [["proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "lot#" in str(response.json()["detail"]).lower()

    async def test_missing_proxy_email_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "proxy email" in str(response.json()["detail"]).lower()

    async def test_invalid_file_type_returns_415(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.pdf", b"garbage", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_building_not_found_returns_404(
        self, client: AsyncClient
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_no_headers_csv_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", b"", "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-proxies (Excel)
# ---------------------------------------------------------------------------


class TestImportProxiesExcel:
    # --- Happy path ---

    async def test_upserts_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 2
        assert data["removed"] == 0

    async def test_blank_proxy_email_removes_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Create proxy first via CSV
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )

        # Remove via Excel
        xlsx = make_excel(["Lot#", "Proxy Email"], [["1A", None]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["removed"] == 1

    async def test_skips_unknown_lot_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [["9Z", "proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["skipped"] == 1
        assert data["upserted"] == 0

    async def test_skips_blank_rows_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"], [None, None]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    # --- Input validation ---

    async def test_missing_headers_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "proxy email" in str(response.json()["detail"]).lower()

    async def test_invalid_excel_file(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", b"not-an-excel-file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "invalid excel" in str(response.json()["detail"]).lower()

    async def test_excel_no_data_rows(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_completely_empty_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """An Excel file with absolutely no rows (not even headers) should return 422."""
        wb = openpyxl.Workbook()
        # Don't write anything to the sheet
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("empty.xlsx", empty_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "no headers" in str(response.json()["detail"]).lower()

    async def test_row_with_fewer_cells_than_headers(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Row that is shorter than the header row — _cell returns '' for out-of-bounds index."""
        # Build an Excel manually where a data row only has 1 cell (Lot#) but not Proxy Email
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Proxy Email"])
        # Write a row with only the first cell populated; second cell omitted
        ws.cell(row=2, column=1, value="1A")
        # Intentionally do NOT set column 2 — openpyxl will yield None for it
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("short.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        # A row with blank proxy email should count as removed/noop (blank -> no existing proxy)
        assert response.status_code == 200
        data = response.json()
        assert data["removed"] == 0  # no existing proxy to remove
        assert data["upserted"] == 0  # blank proxy email


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-financial-positions (CSV)
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsCSV:
    # --- Happy path ---

    async def test_updates_financial_positions(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "In Arrear"], ["2B", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 2
        assert data["skipped"] == 0

        # Verify DB
        result = await db_session.execute(
            select(LotOwner).where(
                LotOwner.building_id == building_with_owners.id,
                LotOwner.lot_number == "1A",
            )
        )
        lo1 = result.scalar_one()
        assert lo1.financial_position == FinancialPosition.in_arrear

    async def test_accepted_value_in_arrear_lowercase(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "in arrear"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_accepted_value_in_arrear_underscore(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "in_arrear"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_accepted_value_normal_uppercase(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "NORMAL"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_unknown_lot_number_skipped(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "Normal"], ["ZZZ", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 1
        assert data["skipped"] == 1

    async def test_lots_not_in_file_unaffected(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Set 1A to in_arrear
        csv_data1 = make_csv(["Lot#", "Financial Position"], [["1A", "In Arrear"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data1, "text/csv")},
        )

        # Second file only updates 2B; 1A should remain in_arrear
        csv_data2 = make_csv(["Lot#", "Financial Position"], [["2B", "Normal"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data2, "text/csv")},
        )

        result = await db_session.execute(
            select(LotOwner).where(
                LotOwner.building_id == building_with_owners.id,
                LotOwner.lot_number == "1A",
            )
        )
        lo1 = result.scalar_one()
        assert lo1.financial_position == FinancialPosition.in_arrear

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position", "Extra"],
            [["1A", "Normal", "ignore"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_empty_csv_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0

    async def test_case_insensitive_headers(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["LOT#", "FINANCIAL POSITION"],
            [["1A", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    # --- Input validation ---

    async def test_invalid_financial_position_value_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "BadValue"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        detail = str(response.json()["detail"])
        assert "BadValue" in detail

    async def test_multiple_invalid_rows_all_listed(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "Wrong"], ["2B", "AlsoBad"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        detail = response.json()["detail"]
        assert isinstance(detail, list)
        assert len(detail) == 2

    async def test_empty_financial_position_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_missing_lot_hash_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Financial Position"], [["Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "lot#" in str(response.json()["detail"]).lower()

    async def test_missing_financial_position_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "financial position" in str(response.json()["detail"]).lower()

    async def test_invalid_file_type_returns_415(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.pdf", b"garbage", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_building_not_found_returns_404(
        self, client: AsyncClient
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_no_headers_csv_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", b"", "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-financial-positions (Excel)
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsExcel:
    # --- Happy path ---

    async def test_updates_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Financial Position"],
            [["1A", "In Arrear"], ["2B", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 2
        assert data["skipped"] == 0

    async def test_invalid_value_in_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [["1A", "Bad"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422

    async def test_skips_unknown_lot_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [["9Z", "Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["skipped"] == 1

    async def test_skips_blank_rows_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Financial Position"],
            [["1A", "Normal"], [None, None]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    # --- Input validation ---

    async def test_missing_headers_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "financial position" in str(response.json()["detail"]).lower()

    async def test_invalid_excel_file(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", b"not-an-excel-file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "invalid excel" in str(response.json()["detail"]).lower()

    async def test_excel_no_data_rows(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0

    async def test_completely_empty_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """An Excel file with absolutely no rows should return 422."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("empty.xlsx", empty_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "no headers" in str(response.json()["detail"]).lower()

    async def test_row_with_fewer_cells_than_headers(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Row shorter than header — _cell returns '' for out-of-bounds, empty fp raises 422."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Financial Position"])
        ws.cell(row=2, column=1, value="1A")
        # column 2 intentionally omitted
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("short.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        # Empty financial_position should be caught as validation error
        assert response.status_code == 422
